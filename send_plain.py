# send_plain.py
import os
import smtplib

# 엑셀 다루기
from openpyxl import load_workbook

# 이메일 메시지를 이진 데이터로 바꿔주는 인코더
from email import encoders

from email.utils import formataddr  # 튜플 만들어주는 모듈
from email.mime.text import MIMEText  # 메일 제목과 본문을 위한 모듈
from email.mime.multipart import MIMEMultipart  # 첨부파일을 위한 모듈

# 텍스트 형식
from email.mime.text import MIMEText
# 이미지 형식
from email.mime.image import MIMEImage
# 오디오 형식
from email.mime.audio import MIMEAudio

# 위의 모든 객체들을 생성할 수 있는 기본 객체
# MIMEBase(_maintype, _subtype)
# MIMEBase(<메인 타입>, <서브 타입>)
from email.mime.base import MIMEBase

msg_dict = {
    # 텍스트 첨부파일
    'text': {'maintype': 'text', 'subtype': 'plain', 'filename': 'test.txt'},
    # 이미지 첨부파일
    'image': {'maintype': 'image', 'subtype': 'jpg', 'filename': 'hacker.png'},
    # 오디오 첨부파일
    # 'audio': {'maintype': 'audio', 'subtype': 'mp3', 'filename': 'test.mp3'},
    # 비디오 첨부파일
    # 'video': {'maintype': 'video', 'subtype': 'mp4', 'filename': 'test.mp4'},
    # 그외 첨부파일
    'application': {'maintype': 'application', 'subtype': 'octect-stream', 'filename': 'test.pdf'}
}

reciever = {}

# 엑셀에서 메일 읽어오기
wb = load_workbook('reciever.xlsx', data_only=True, read_only=True)
ws = wb.active

for col in ws.iter_rows():
    index = 0
    for cell in col:
        if(index == 0):
            # 첫번째 나오는아이 -> 이름 (엑셀에서 이름, 메일 순서로 받아옴) -> 딕셔너리 key로 만듬
            tmp_key = cell.value
        else:
            # 두번째 나오는 아이 -> 메일 -> 딕셔너리에 value로 들어감
            reciever[tmp_key] = cell.value
        print(cell.value)
        index = index+1

print(reciever)


def make_multimsg(msg_dict):
    multi = MIMEMultipart(_subtype='mixed')

    for key, value in msg_dict.items():
        # 각 타입에 적절한 MIMExxx()함수를 호출하여 msg 객체를 생성한다.
        if key == 'text':
            with open(value['filename'], encoding='utf-8') as fp:
                msg = MIMEText(fp.read(), _subtype=value['subtype'])
        elif key == 'image':
            with open(value['filename'], 'rb') as fp:
                msg = MIMEImage(fp.read(), _subtype=value['subtype'])
        elif key == 'audio':
            with open(value['filename'], 'rb') as fp:
                msg = MIMEAudio(fp.read(), _subtype=value['subtype'])
        else:
            with open(value['filename'], 'rb') as fp:
                msg = MIMEBase(value['maintype'],  _subtype=value['subtype'])
                msg.set_payload(fp.read())
                encoders.encode_base64(msg)
        # 파일 이름을 첨부파일 제목으로 추가
        msg.add_header('Content-Disposition', 'attachment',
                       filename=value['filename'])
        # 첨부파일 multi에 추가
        multi.attach(msg)
    # multi를 반환, 이걸 message에 attach를 통해 합쳐줌
    return multi


for name, email in reciever.items():
    try:
        session = None
        # SMTP 세션 생성
        session = smtplib.SMTP('smtp.gmail.com', 587)
        # debug 내용 보고 싶으면 True로
        session.set_debuglevel(False)

        # SMTP 계정 인증 설정
        session.ehlo()
        session.starttls()
        session.login('gygh7562@gmail.com', 'wbfrncdwtwqxzwrz')

        # 보내는 사람('이름', '메일 주소')
        from_addr = formataddr(('UN사무총장', 'test@kshield.com'))
        to_addr = formataddr((name, email))
        print("from_addr : "+from_addr)
        print("to_addr : "+to_addr)

        # 메일 콘텐츠 설정
        message = MIMEMultipart("alternative")
        message.set_charset('utf-8')

        # 메일 송/수신 옵션 설정
        message['From'] = from_addr
        message['To'] = to_addr
        message['Subject'] = '안녕하세요 보안방역반입니다.'

        # 첨부파일 넣기
        multi = make_multimsg(msg_dict)
        message.attach(multi)

        # 메일 콘텐츠 - 내용
        body = '''
                <h1>hi juntheworld</h1>
                <a href = "www.facebook.com">go to facebook</a>
                '''
        bodyPart = MIMEText(body, 'html', 'utf-8')
        message.attach(bodyPart)

        # 메일 발송
        session.sendmail("jun", to_addr, message.as_string())
        print('Successfully sent the mail!!!')

    except Exception as e:
        print(e)
    finally:
        if session is not None:
            session.quit()
