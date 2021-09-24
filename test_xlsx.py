# send_plain.py
import os
import smtplib

# 엑셀 다루기
from openpyxl import load_workbook

wb = load_workbook('reciever.xlsx', data_only=True, read_only=True)
ws = wb.active

for 행 in ws.iter_rows():
    index = 0
    for cell in 행:
        if(index == 1):
            print("이거 메일주소")
        else:
            print("이거 이름")
        print(cell.value)
        index = index+1

# print(ws["A1"].value)
#print(ws.cell(1, 2).value)
